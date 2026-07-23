"""
Invoice Chaser — SaaS financial model builder.
Cohort-driven, 36-month monthly model with scenario switching.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import LineChart, Reference

N = 36           # months
C0 = 3           # first month column index (C)

def mc(t):
    """month t (1-based) -> column letter"""
    return get_column_letter(C0 + t - 1)

LAST = mc(N)

# ---------------------------------------------------------------- palette
NAVY  = "1F3864"
TEAL  = "31859C"
LTEAL = "DEEBF0"
GREY  = "F2F2F2"

F_TITLE   = Font(name="Calibri", size=14, bold=True, color=NAVY)
F_SUB     = Font(name="Calibri", size=10, italic=True, color="595959")
F_SEC     = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
F_HDR     = Font(name="Calibri", size=10, bold=True, color=NAVY)
F_LBL     = Font(name="Calibri", size=10, color="000000")
F_LBLB    = Font(name="Calibri", size=10, bold=True, color=NAVY)
F_IN      = Font(name="Calibri", size=10, color="0000FF")
F_CALC    = Font(name="Calibri", size=10, color="000000")
F_LINK    = Font(name="Calibri", size=10, color="008000")
F_NOTE    = Font(name="Calibri", size=9, italic=True, color="7F7F7F")
COLOR_OF  = {}

COLOR_OF[F_CALC] = "000000"
COLOR_OF[F_LINK] = "008000"
COLOR_OF[F_NOTE] = "7F7F7F"
COLOR_OF[F_IN] = "0000FF"

FILL_SEC  = PatternFill("solid", fgColor=NAVY)
FILL_KEY  = PatternFill("solid", fgColor="FFFF00")
FILL_BAND = PatternFill("solid", fgColor=LTEAL)
FILL_GREY = PatternFill("solid", fgColor=GREY)

TOP = Border(top=Side(style="thin", color=NAVY))

EUR0 = '€#,##0;(€#,##0);-'
EUR2 = '€#,##0.00;(€#,##0.00);-'
EUR3 = '€#,##0.000;(€#,##0.000);-'
PCT1 = '0.0%;(0.0%);-'
PCT2 = '0.00%;(0.00%);-'
NUM0 = '#,##0;(#,##0);-'
NUM1 = '#,##0.0;(#,##0.0);-'
MULT = '0.0x'
IDX  = '0'

wb = Workbook()

# ================================================================ READ ME
ws = wb.active
ws.title = "Read Me"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 34
ws.column_dimensions["C"].width = 96

r = 2
ws.cell(r, 2, "INVOICE CHASER — SaaS Financial Model").font = F_TITLE
r += 1
ws.cell(r, 2, "Automated invoice follow-up for French auto-entrepreneurs · 36-month monthly build").font = F_SUB
r += 2

def readme_sec(title):
    global r
    ws.cell(r, 2, title).font = F_SEC
    ws.cell(r, 2).fill = FILL_SEC
    ws.cell(r, 3).fill = FILL_SEC
    r += 1

def readme_row(k, v):
    global r
    ws.cell(r, 2, k).font = F_LBLB
    c = ws.cell(r, 3, v)
    c.font = F_LBL
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = max(15, 13 * (len(v) // 92 + 1))
    r += 1

readme_sec("HOW TO USE THIS MODEL")
readme_row("Change one cell to re-run",
           "Assumptions!B5 is the scenario switch. Enter 1 (Conservative), 2 (Base) or 3 (Aggressive). "
           "Every sheet recalculates from it. Example: set B5 = 3 and the Dashboard repoints to the Aggressive case.")
readme_row("Change the drivers",
           "On Assumptions, edit columns C, D and E (the three scenario columns). Column F is the LIVE column — "
           "it is a formula, do not type over it. Everything downstream reads column F only.")
readme_row("Nothing else is an input",
           "Cohorts, P&L, Cash Flow, Unit Economics and Dashboard are 100% formula-driven. If you need to change a "
           "number, change it on Assumptions.")
r += 1

readme_sec("COLOUR LEGEND")
readme_row("Blue text", "Hardcoded input. Safe to edit.")
readme_row("Yellow fill", "Key lever — the cells worth stress-testing first.")
readme_row("Black text", "Formula. Do not overwrite.")
readme_row("Green text", "Link to another sheet. Do not overwrite.")
r += 1

readme_sec("SHEET MAP")
readme_row("Assumptions", "All drivers, three scenarios side by side, plus a derived block (ARPU, blended churn, implied lifetime).")
readme_row("Cohorts", "Acquisition funnel, then two 36x36 cohort waterfalls — one for monthly plans (constant monthly churn), "
                      "one for annual plans (no churn inside the committed year, renewal decision at each 12-month anniversary).")
readme_row("P&L", "Billings vs recognised revenue, cost of revenue, gross margin, opex, EBITDA.")
readme_row("Cash Flow", "Cash walk plus a deferred revenue rollforward for annual prepayments, and an integrity check row.")
readme_row("Unit Economics", "LTV, CAC (paid and blended), payback, LTV:CAC, and two sensitivity grids.")
readme_row("Dashboard", "Headline outputs, breakeven months, funding requirement, NPV, and two charts.")
r += 1

readme_sec("HOW THE ANNUAL PLAN IS MODELLED")
readme_row("Why it matters", "Annual prepayment is the single biggest cash lever a low-ticket SaaS has. Ignoring it overstates "
                             "the funding gap; treating cash as revenue overstates the P&L. This model separates the two.")
readme_row("Billings", "An annual customer is charged 10x the monthly price upfront (2 months free) in their cohort month, "
                       "and again at each anniversary if they renew.")
readme_row("Revenue", "Recognised straight-line: 10/12 of the monthly price, each month, for 12 months.")
readme_row("The difference", "Sits in deferred revenue on the Cash Flow sheet. Cash Flow row 22 checks that "
                             "Net cash flow − EBITDA − Δdeferred = 0 in every month. It must read zero across the row.")
r += 1

readme_sec("HEALTH WARNING ON THE NUMBERS")
readme_row("These are benchmarks, not evidence",
           "Every operating assumption here is a placeholder drawn from typical low-ticket, self-serve B2B SaaS ranges. "
           "None of it is observed data for this product. The model is a decision framework, not a forecast.")
readme_row("The three that decide the business",
           "Monthly churn, trial-to-paid conversion, and blended CPC. The Unit Economics sensitivity grids show that the "
           "verdict flips inside plausible ranges of these three — so they are what validation has to resolve first.")
readme_row("Cheapest way to test them",
           "Churn and conversion need a landing page and ~20 customer conversations, not a spreadsheet. CPC can be read off "
           "a €200 test campaign in a week. Do that before trusting any output on the Dashboard.")
readme_row("Not modelled", "VAT/TVA treatment, corporate tax, any founder opportunity cost beyond the salary line, "
                           "and competitive response. Add them before this goes in front of anyone external.")

# ================================================================ ASSUMPTIONS
aws = wb.create_sheet("Assumptions")
aws.sheet_view.showGridLines = False
aws.column_dimensions["A"].width = 46
aws.column_dimensions["B"].width = 17
for col in "CDEF":
    aws.column_dimensions[col].width = 13
aws.column_dimensions["G"].width = 78

aws.cell(2, 1, "ASSUMPTIONS & SCENARIO DRIVERS").font = F_TITLE
aws.cell(3, 1, "Edit columns C–E only. Column F is the live selection and feeds every other sheet.").font = F_SUB

aws.cell(5, 1, "Scenario switch  (1 = Conservative, 2 = Base, 3 = Aggressive)").font = F_LBLB
sw = aws.cell(5, 2, 2)
sw.font = F_IN
sw.fill = FILL_KEY
sw.number_format = IDX
sw.alignment = Alignment(horizontal="center")
sw.border = Border(*[Side(style="thin", color=NAVY)] * 4)

dv = DataValidation(type="whole", operator="between", formula1=1, formula2=3,
                    showErrorMessage=True, errorTitle="Scenario",
                    error="Enter 1 (Conservative), 2 (Base) or 3 (Aggressive).")
aws.add_data_validation(dv)
dv.add(sw)

aws.cell(6, 1, "Active scenario").font = F_LBLB
aws.cell(6, 2, "=INDEX(C6:E6,$B$5)").font = F_CALC
aws.cell(6, 3, "Conservative").font = F_IN
aws.cell(6, 4, "Base").font = F_IN
aws.cell(6, 5, "Aggressive").font = F_IN
for c in range(3, 6):
    aws.cell(6, c).alignment = Alignment(horizontal="center")

HDR_ROW = 8
for c, h in [(1, "Driver"), (2, "Unit"), (3, "Conservative"), (4, "Base"),
             (5, "Aggressive"), (6, "LIVE"), (7, "Basis / how to validate")]:
    cell = aws.cell(HDR_ROW, c, h)
    cell.font = F_SEC
    cell.fill = FILL_SEC
    cell.alignment = Alignment(horizontal="center" if c > 1 else "left")

# (kind, key, label, unit, cons, base, aggr, fmt, key_lever, note)
S = "sec"; I = "in"; D = "der"

ROWS = [
    (S, None, "PRICING & PLANS", None, None, None, None, None, False, None),
    (I, "price_solo", "Solo plan price", "€ / month", 9, 9, 12, EUR2, False,
     "Placeholder. Anchored below what a comptable charges for the same chasing. Validate with pricing interviews."),
    (I, "price_pro", "Pro plan price", "€ / month", 19, 19, 24, EUR2, False,
     "Placeholder. Pro = multi-client, custom sequences, reminders by SMS."),
    (I, "pro_mix", "Pro plan share of customers", "% of customers", 0.20, 0.30, 0.35, PCT1, False,
     "Placeholder. Share of base on the higher tier."),
    (I, "annual_share", "Annual plan share of new customers", "% of new", 0.15, 0.20, 0.25, PCT1, True,
     "Key cash lever. Self-serve SaaS at this price point rarely pushes past ~25% annual without a hard discount."),
    (I, "annual_free", "Annual plan — months free vs 12x monthly", "months", 2, 2, 2, NUM0, False,
     "Standard 2-months-free annual offer, i.e. pay for 10, get 12."),

    (S, None, "ACQUISITION", None, None, None, None, None, False, None),
    (I, "spend_m1", "Paid marketing spend, month 1", "€ / month", 300, 500, 900, EUR0, False,
     "Founder-funded test budget in month 1."),
    (I, "spend_growth", "Paid spend growth", "% / month", 0.06, 0.08, 0.10, PCT1, False,
     "Reinvestment pace. Deliberately slow — spend is capped below."),
    (I, "spend_cap", "Paid spend ceiling", "€ / month", 3000, 6000, 12000, EUR0, False,
     "Ceiling. Reflects a genuinely narrow niche — French micro-entrepreneurs searching for invoice chasing. "
     "The addressable paid inventory runs out."),
    (I, "cpc", "Blended cost per click", "€", 1.10, 0.80, 0.60, EUR2, True,
     "Key lever. Placeholder for French-language Google/Meta on low-competition long-tail terms. "
     "Testable for ~€200 in one week — do this first."),
    (I, "signup_rate", "Landing page → free trial", "%", 0.025, 0.040, 0.055, PCT1, False,
     "Placeholder. Typical range for a self-serve tool with a clear single job-to-be-done."),
    (I, "trial_conv", "Free trial → paid", "%", 0.15, 0.22, 0.28, PCT1, True,
     "Key lever. Trial-to-paid for no-credit-card self-serve trials. Drives CAC almost as hard as CPC."),
    (I, "org_m1", "Organic trials, month 1", "trials", 10, 20, 35, NUM0, False,
     "Content/SEO and community (forums, LinkedIn, auto-entrepreneur groups). Zero marginal cost, slow to build."),
    (I, "org_growth", "Organic trial growth", "% / month", 0.08, 0.12, 0.16, PCT1, False,
     "Compounding content effect."),
    (I, "org_cap", "Organic trial ceiling", "trials / month", 150, 400, 800, NUM0, False,
     "Ceiling. Organic demand for this niche is finite."),
    (I, "ref_rate", "Referral trials per active customer", "/ month", 0.005, 0.015, 0.030, PCT2, False,
     "Word of mouth. Freelancers talk, but they are not a dense network."),

    (S, None, "RETENTION", None, None, None, None, None, False, None),
    (I, "churn_monthly", "Monthly-plan gross churn", "% / month", 0.080, 0.060, 0.045, PCT1, True,
     "Single most important assumption. High by SaaS standards because the customer base itself churns — "
     "micro-entrepreneurs deregister, go salaried, or stop invoicing. This is not a product-quality number."),
    (I, "renewal_annual", "Annual-plan renewal rate at each anniversary", "%", 0.60, 0.70, 0.80, PCT1, False,
     "Applied once every 12 months. Annual customers cannot churn mid-commitment."),

    (S, None, "COST OF REVENUE", None, None, None, None, None, False, None),
    (I, "cost_infra", "Infrastructure & automation per customer", "€ / cust / month", 0.45, 0.35, 0.28, EUR3, False,
     "n8n execution volume, storage, API calls. Falls with scale as fixed capacity is absorbed."),
    (I, "cost_email", "Email & messaging per customer", "€ / cust / month", 0.15, 0.10, 0.08, EUR3, False,
     "Transactional email volume — each customer sends multiple reminder sequences per month."),
    (I, "cost_support", "Support cost per customer", "€ / cust / month", 0.50, 0.30, 0.20, EUR3, False,
     "Founder time costed at a notional rate. At €9/month, support is what kills the margin — watch this line."),
    (I, "pay_pct", "Payment processing — % of billings", "%", 0.015, 0.015, 0.015, PCT1, False,
     "Stripe standard European card rate. Verify against current Stripe pricing before relying on it."),
    (I, "pay_fixed", "Payment processing — fee per transaction", "€", 0.25, 0.25, 0.25, EUR2, False,
     "Stripe per-transaction fee. This is why annual billing helps margin, not just cash — one charge, not twelve."),
    (I, "cost_fixed", "Fixed platform cost", "€ / month", 90, 70, 70, EUR0, False,
     "n8n cloud plan, hosting baseline, domain. Does not scale with customers."),

    (S, None, "OPERATING EXPENSES", None, None, None, None, None, False, None),
    (I, "founder_pay", "Founder compensation once drawn", "€ / month", 1500, 2500, 3500, EUR0, False,
     "Set by you. Note this is a side project — the real cost is hours, not euros, until this line switches on."),
    (I, "founder_start", "First month founder is paid", "month #", 13, 13, 10, IDX, True,
     "Key lever. The month the business stops being free to run. Moving this earlier is the fastest way to break the cash line."),
    (I, "opex_tools", "Tools & software", "€ / month", 150, 120, 120, EUR0, False,
     "Stripe dashboard extras, analytics, design, hosting for the marketing site."),
    (I, "opex_admin", "Legal, accounting & admin", "€ / month", 120, 80, 80, EUR0, False,
     "Accountant, company admin, CGU/RGPD review. RGPD matters here — the product touches client contact data."),
    (I, "opex_infl", "Fixed opex inflation", "% / month", 0.005, 0.005, 0.005, PCT2, False,
     "Applied to tools and admin only. ~6% annualised."),

    (S, None, "FINANCING & VALUATION", None, None, None, None, None, False, None),
    (I, "cash_open", "Opening cash (founder contribution)", "€", 5000, 5000, 5000, EUR0, False,
     "Set by you. The model tells you whether this is enough."),
    (I, "disc_annual", "Discount rate", "% / year", 0.30, 0.25, 0.20, PCT1, False,
     "Pre-revenue, single-founder, unvalidated demand. High rate is doing real work here, not decoration."),

    (S, None, "DERIVED  (calculated — do not edit)", None, None, None, None, None, False, None),
    (D, "arpu_monthly", "Blended list price — monthly plan", "€ / month",
     "={price_solo}*(1-{pro_mix})+{price_pro}*{pro_mix}", None, None, EUR2, False,
     "Weighted by Pro mix."),
    (D, "arpu_annual_eff", "Effective ARPU — annual plan", "€ / month",
     "={arpu_monthly}*(12-{annual_free})/12", None, None, EUR2, False,
     "Recognised straight-line over 12 months after the 2-months-free discount."),
    (D, "arpu_blended", "Blended ARPU — all plans", "€ / month",
     "={arpu_monthly}*(1-{annual_share})+{arpu_annual_eff}*{annual_share}", None, None, EUR2, False,
     "Used for LTV."),
    (D, "annual_upfront", "Annual plan — cash charged upfront", "€",
     "={arpu_monthly}*(12-{annual_free})", None, None, EUR0, False,
     "Billed in the cohort month and at each renewal."),
    (D, "churn_annual_eq", "Annual plan — monthly-equivalent churn", "% / month",
     "=1-{renewal_annual}^(1/12)", None, None, PCT2, False,
     "Anniversary renewal rate expressed as an equivalent constant monthly rate, for the blend below."),
    (D, "churn_blended", "Blended monthly churn", "% / month",
     "={churn_monthly}*(1-{annual_share})+{churn_annual_eq}*{annual_share}", None, None, PCT2, False,
     "Weighted by plan mix. This is the churn that drives LTV."),
    (D, "life_months", "Implied average customer lifetime", "months",
     "=IF({churn_blended}=0,0,1/{churn_blended})", None, None, NUM1, False,
     "1 / blended churn."),
    (D, "paid_cac_theo", "Paid CAC (from funnel maths)", "€",
     "=IF({signup_rate}*{trial_conv}=0,0,{cpc}/({signup_rate}*{trial_conv}))", None, None, EUR2, False,
     "CPC / (signup rate x trial-to-paid). Cost to buy one paying customer through paid channels."),
    (D, "disc_monthly", "Monthly discount rate", "% / month",
     "=(1+{disc_annual})^(1/12)-1", None, None, PCT2, False,
     "Annual rate de-annualised for the monthly NPV."),
]

AREF = {}
r = HDR_ROW + 1
_pending = []
for item in ROWS:
    kind = item[0]
    if kind == S:
        for c in range(1, 8):
            cell = aws.cell(r, c, item[2] if c == 1 else None)
            cell.fill = FILL_BAND
            if c == 1:
                cell.font = F_HDR
        r += 1
        continue
    key, label, unit = item[1], item[2], item[3]
    AREF[key] = f"Assumptions!$F${r}"
    aws.cell(r, 1, label).font = F_LBL
    aws.cell(r, 2, unit).font = F_NOTE
    aws.cell(r, 2).alignment = Alignment(horizontal="center")
    fmt, is_key, note = item[7], item[8], item[9]
    if kind == I:
        for c, v in zip((3, 4, 5), (item[4], item[5], item[6])):
            cell = aws.cell(r, c, v)
            cell.font = F_IN
            cell.number_format = fmt
            if is_key:
                cell.fill = FILL_KEY
        f = aws.cell(r, 6, f"=INDEX(C{r}:E{r},$B$5)")
        f.font = F_CALC
        f.number_format = fmt
        f.fill = FILL_GREY
    else:
        _pending.append((r, item[4], fmt))
    nc = aws.cell(r, 7, note)
    nc.font = F_NOTE
    nc.alignment = Alignment(wrap_text=True, vertical="top")
    aws.row_dimensions[r].height = max(15, 11 * (len(note) // 88 + 1))
    r += 1

def A(key):
    return AREF[key]

def sub(tpl):
    out = tpl
    for k, v in AREF.items():
        out = out.replace("{" + k + "}", v)
    return out

for rr, tpl, fmt in _pending:
    cell = aws.cell(rr, 6, sub(tpl))
    cell.font = F_CALC
    cell.number_format = fmt
    cell.fill = FILL_GREY

aws.freeze_panes = "A9"

# ================================================================ shared helpers
def month_header(sheet, title, subtitle):
    sheet.sheet_view.showGridLines = False
    sheet.column_dimensions["A"].width = 44
    sheet.column_dimensions["B"].width = 15
    for t in range(1, N + 1):
        sheet.column_dimensions[mc(t)].width = 11
    sheet.cell(2, 1, title).font = F_TITLE
    sheet.cell(3, 1, subtitle).font = F_SUB
    sheet.cell(5, 1, "Month index").font = F_HDR
    sheet.cell(5, 2, "#").font = F_NOTE
    sheet.cell(5, 2).alignment = Alignment(horizontal="center")
    for t in range(1, N + 1):
        c = sheet.cell(5, C0 + t - 1, t)
        c.font = F_HDR
        c.number_format = IDX
        c.alignment = Alignment(horizontal="center")
        c.fill = FILL_BAND
    sheet.freeze_panes = "C6"

def section(sheet, row, title):
    for c in range(1, C0 + N):
        cell = sheet.cell(row, c, title if c == 1 else None)
        cell.fill = FILL_SEC
        if c == 1:
            cell.font = F_SEC

def band(sheet, row, title):
    for c in range(1, C0 + N):
        cell = sheet.cell(row, c, title if c == 1 else None)
        cell.fill = FILL_BAND
        if c == 1:
            cell.font = F_HDR

def line(sheet, row, label, unit, fn, fmt, font=F_CALC, bold=False, rule=False):
    lc = sheet.cell(row, 1, label)
    lc.font = F_LBLB if bold else F_LBL
    uc = sheet.cell(row, 2, unit)
    uc.font = F_NOTE
    uc.alignment = Alignment(horizontal="center")
    for t in range(1, N + 1):
        cell = sheet.cell(row, C0 + t - 1, sub(fn(t)))
        cell.font = Font(name="Calibri", size=10, bold=bold, color=COLOR_OF[font])
        cell.number_format = fmt
        if rule:
            cell.border = TOP
    if rule:
        lc.border = TOP
        uc.border = TOP

# ================================================================ COHORTS
cws = wb.create_sheet("Cohorts")
month_header(cws, "COHORT ENGINE",
             "Acquisition funnel, then customer waterfalls by cohort. Monthly plans churn every month; "
             "annual plans are locked for 12 and face a renewal decision at each anniversary.")

R_SPEND, R_VIS, R_PTRI, R_OTRI, R_RTRI, R_TTRI, R_NEW, R_NEWM, R_NEWA = 8, 9, 10, 11, 12, 13, 14, 15, 16
R_ACTM, R_ACTA, R_ACT, R_OPEN, R_ABILL = 19, 20, 21, 22, 23
GA_HDR = 26
GA0 = GA_HDR + 1
GB_HDR = GA0 + N + 2
GB0 = GB_HDR + 1

section(cws, 7, "ACQUISITION FUNNEL")
line(cws, R_SPEND, "Paid marketing spend", "€",
     lambda t: f"=MIN({A('spend_cap')},{A('spend_m1')}*(1+{A('spend_growth')})^({mc(t)}5-1))", EUR0)
line(cws, R_VIS, "Paid visitors", "#",
     lambda t: f"=IF({A('cpc')}=0,0,{mc(t)}{R_SPEND}/{A('cpc')})", NUM0)
line(cws, R_PTRI, "Paid trials", "#",
     lambda t: f"={mc(t)}{R_VIS}*{A('signup_rate')}", NUM1)
line(cws, R_OTRI, "Organic trials", "#",
     lambda t: f"=MIN({A('org_cap')},{A('org_m1')}*(1+{A('org_growth')})^({mc(t)}5-1))", NUM1)
line(cws, R_RTRI, "Referral trials", "#",
     lambda t: f"={mc(t)}{R_OPEN}*{A('ref_rate')}", NUM1)
line(cws, R_TTRI, "Total trials", "#",
     lambda t: f"=SUM({mc(t)}{R_PTRI}:{mc(t)}{R_RTRI})", NUM1, bold=True, rule=True)
line(cws, R_NEW, "New paying customers", "#",
     lambda t: f"={mc(t)}{R_TTRI}*{A('trial_conv')}", NUM1, bold=True)
line(cws, R_NEWM, "    — new on monthly plan", "#",
     lambda t: f"={mc(t)}{R_NEW}*(1-{A('annual_share')})", NUM1)
line(cws, R_NEWA, "    — new on annual plan", "#",
     lambda t: f"={mc(t)}{R_NEW}*{A('annual_share')}", NUM1)

section(cws, 18, "ACTIVE CUSTOMERS")
line(cws, R_ACTM, "Monthly-plan customers (closing)", "#",
     lambda t: f"=SUM({mc(t)}{GA0}:{mc(t)}{GA0+N-1})", NUM1, font=F_CALC)
line(cws, R_ACTA, "Annual-plan customers (closing)", "#",
     lambda t: f"=SUM({mc(t)}{GB0}:{mc(t)}{GB0+N-1})", NUM1, font=F_CALC)
line(cws, R_ACT, "Total active customers (closing)", "#",
     lambda t: f"={mc(t)}{R_ACTM}+{mc(t)}{R_ACTA}", NUM1, bold=True, rule=True)
line(cws, R_OPEN, "Total active customers (opening)", "#",
     lambda t: "=0" if t == 1 else f"={mc(t-1)}{R_ACT}", NUM1)

def abill(t):
    parts = [f"{mc(t)}{R_NEWA}"]
    if t > 12:
        parts.append(f"{mc(t)}{GB0 + (t - 12) - 1}")
    if t > 24:
        parts.append(f"{mc(t)}{GB0 + (t - 24) - 1}")
    return "=" + "+".join(parts)

line(cws, R_ABILL, "Annual plans billed this month", "#", abill, NUM1)
cws.cell(R_ABILL, 1).font = F_LBL
cws.cell(R_ABILL + 1, 1, "New annual sign-ups plus the cohorts hitting their 12- and 24-month anniversary "
                          "(post-renewal count).").font = F_NOTE

# --- Grid A: monthly plan
band(cws, GA_HDR - 1, "COHORT WATERFALL — MONTHLY PLAN  (active customers)")
cws.cell(GA_HDR, 1, "Cohort ↓ / Month →").font = F_HDR
cws.cell(GA_HDR, 2, "#").font = F_NOTE
for t in range(1, N + 1):
    c = cws.cell(GA_HDR, C0 + t - 1, t)
    c.font = F_HDR
    c.number_format = IDX
    c.alignment = Alignment(horizontal="center")

for m in range(1, N + 1):
    row = GA0 + m - 1
    cws.cell(row, 1, f"Cohort M{m}").font = F_LBL
    cws.cell(row, 2, "#").font = F_NOTE
    cws.cell(row, 2).alignment = Alignment(horizontal="center")
    for t in range(1, N + 1):
        cell = cws.cell(row, C0 + t - 1)
        cell.number_format = NUM1
        if t < m:
            cell.value = 0
            cell.font = F_NOTE
        elif t == m:
            cell.value = f"={mc(t)}{R_NEWM}"
            cell.font = F_CALC
        else:
            cell.value = sub(f"={mc(t-1)}{row}*(1-{A('churn_monthly')})")
            cell.font = F_CALC

# --- Grid B: annual plan
band(cws, GB_HDR - 1, "COHORT WATERFALL — ANNUAL PLAN  (active customers)")
cws.cell(GB_HDR, 1, "Cohort ↓ / Month →").font = F_HDR
cws.cell(GB_HDR, 2, "#").font = F_NOTE
for t in range(1, N + 1):
    c = cws.cell(GB_HDR, C0 + t - 1, t)
    c.font = F_HDR
    c.number_format = IDX
    c.alignment = Alignment(horizontal="center")

for m in range(1, N + 1):
    row = GB0 + m - 1
    cws.cell(row, 1, f"Cohort M{m}").font = F_LBL
    cws.cell(row, 2, "#").font = F_NOTE
    cws.cell(row, 2).alignment = Alignment(horizontal="center")
    for t in range(1, N + 1):
        cell = cws.cell(row, C0 + t - 1)
        cell.number_format = NUM1
        age = t - m
        if t < m:
            cell.value = 0
            cell.font = F_NOTE
        elif t == m:
            cell.value = f"={mc(t)}{R_NEWA}"
            cell.font = F_CALC
        elif age % 12 == 0:
            cell.value = sub(f"={mc(t-1)}{row}*{A('renewal_annual')}")
            cell.font = F_CALC
        else:
            cell.value = f"={mc(t-1)}{row}"
            cell.font = F_CALC

# ================================================================ P&L
pws = wb.create_sheet("P&L")
month_header(pws, "PROFIT & LOSS",
             "Billings are cash charged. Revenue is recognised. They differ by the annual prepayment — "
             "the gap is tracked on Cash Flow.")

P_CBM, P_CBA, P_BM, P_BA, P_BT, P_TX = 8, 9, 10, 11, 12, 13
P_RM, P_RA, P_RT, P_ARR, P_GRW = 16, 17, 18, 19, 20
P_CI, P_CE, P_CS, P_CP, P_CF, P_CT = 23, 24, 25, 26, 27, 28
P_GP, P_GM = 30, 31
P_OM, P_OF, P_OT, P_OA, P_OTT = 34, 35, 36, 37, 38
P_EB, P_EBM, P_CUM, P_FLAG = 40, 41, 42, 44

section(pws, 7, "BILLINGS  (cash charged to the customer)")
line(pws, P_CBM, "Monthly-plan customers billed", "#", lambda t: f"=Cohorts!{mc(t)}{R_ACTM}", NUM1, font=F_LINK)
line(pws, P_CBA, "Annual plans billed", "#", lambda t: f"=Cohorts!{mc(t)}{R_ABILL}", NUM1, font=F_LINK)
line(pws, P_BM, "Monthly-plan billings", "€", lambda t: f"={mc(t)}{P_CBM}*{A('arpu_monthly')}", EUR0)
line(pws, P_BA, "Annual-plan billings", "€", lambda t: f"={mc(t)}{P_CBA}*{A('annual_upfront')}", EUR0)
line(pws, P_BT, "Total billings", "€", lambda t: f"={mc(t)}{P_BM}+{mc(t)}{P_BA}", EUR0, bold=True, rule=True)
line(pws, P_TX, "Card transactions", "#", lambda t: f"={mc(t)}{P_CBM}+{mc(t)}{P_CBA}", NUM1)

section(pws, 15, "REVENUE  (recognised)")
line(pws, P_RM, "Monthly-plan subscription revenue", "€",
     lambda t: f"=Cohorts!{mc(t)}{R_ACTM}*{A('arpu_monthly')}", EUR0)
line(pws, P_RA, "Annual-plan subscription revenue", "€",
     lambda t: f"=Cohorts!{mc(t)}{R_ACTA}*{A('arpu_annual_eff')}", EUR0)
line(pws, P_RT, "Total revenue  (= MRR)", "€",
     lambda t: f"={mc(t)}{P_RM}+{mc(t)}{P_RA}", EUR0, bold=True, rule=True)
line(pws, P_ARR, "ARR (run-rate)", "€", lambda t: f"={mc(t)}{P_RT}*12", EUR0)
line(pws, P_GRW, "Revenue growth MoM", "%",
     lambda t: "=0" if t == 1 else f"=IF({mc(t-1)}{P_RT}=0,0,{mc(t)}{P_RT}/{mc(t-1)}{P_RT}-1)", PCT1)

section(pws, 22, "COST OF REVENUE")
line(pws, P_CI, "Infrastructure & automation", "€",
     lambda t: f"=Cohorts!{mc(t)}{R_ACT}*{A('cost_infra')}", EUR0)
line(pws, P_CE, "Email & messaging", "€",
     lambda t: f"=Cohorts!{mc(t)}{R_ACT}*{A('cost_email')}", EUR0)
line(pws, P_CS, "Customer support", "€",
     lambda t: f"=Cohorts!{mc(t)}{R_ACT}*{A('cost_support')}", EUR0)
line(pws, P_CP, "Payment processing", "€",
     lambda t: f"={mc(t)}{P_BT}*{A('pay_pct')}+{mc(t)}{P_TX}*{A('pay_fixed')}", EUR0)
line(pws, P_CF, "Fixed platform cost", "€", lambda t: f"={A('cost_fixed')}", EUR0)
line(pws, P_CT, "Total cost of revenue", "€",
     lambda t: f"=SUM({mc(t)}{P_CI}:{mc(t)}{P_CF})", EUR0, bold=True, rule=True)
line(pws, P_GP, "Gross profit", "€", lambda t: f"={mc(t)}{P_RT}-{mc(t)}{P_CT}", EUR0, bold=True)
line(pws, P_GM, "Gross margin", "%",
     lambda t: f"=IF({mc(t)}{P_RT}=0,0,{mc(t)}{P_GP}/{mc(t)}{P_RT})", PCT1, bold=True)

section(pws, 33, "OPERATING EXPENSES")
line(pws, P_OM, "Paid marketing", "€", lambda t: f"=Cohorts!{mc(t)}{R_SPEND}", EUR0, font=F_LINK)
line(pws, P_OF, "Founder compensation", "€",
     lambda t: f"=IF({mc(t)}5>={A('founder_start')},{A('founder_pay')},0)", EUR0)
line(pws, P_OT, "Tools & software", "€",
     lambda t: f"={A('opex_tools')}*(1+{A('opex_infl')})^({mc(t)}5-1)", EUR0)
line(pws, P_OA, "Legal, accounting & admin", "€",
     lambda t: f"={A('opex_admin')}*(1+{A('opex_infl')})^({mc(t)}5-1)", EUR0)
line(pws, P_OTT, "Total operating expenses", "€",
     lambda t: f"=SUM({mc(t)}{P_OM}:{mc(t)}{P_OA})", EUR0, bold=True, rule=True)

line(pws, P_EB, "EBITDA", "€", lambda t: f"={mc(t)}{P_GP}-{mc(t)}{P_OTT}", EUR0, bold=True, rule=True)
line(pws, P_EBM, "EBITDA margin", "%",
     lambda t: f"=IF({mc(t)}{P_RT}=0,0,{mc(t)}{P_EB}/{mc(t)}{P_RT})", PCT1)
line(pws, P_CUM, "Cumulative EBITDA", "€",
     lambda t: f"={mc(t)}{P_EB}" if t == 1 else f"={mc(t-1)}{P_CUM}+{mc(t)}{P_EB}", EUR0, bold=True)

pws.cell(P_FLAG - 1, 1, "MEMO").font = F_HDR
line(pws, P_FLAG, "Month flag — EBITDA positive", "#",
     lambda t: f"=IF({mc(t)}{P_EB}>0,{mc(t)}5,9999)", IDX, font=F_NOTE)
pws.cell(P_FLAG + 1, 1, "9999 = not yet positive. Dashboard takes the MIN of this row.").font = F_NOTE

# ================================================================ CASH FLOW
fws = wb.create_sheet("Cash Flow")
month_header(fws, "CASH FLOW & DEFERRED REVENUE",
             "Card payments are collected on charge, so collections equal billings. The gap between cash and "
             "EBITDA is the annual prepayment, and it is checked every month.")

C_OP, C_COL, C_COGS, C_OPEX, C_NET, C_CL = 8, 9, 10, 11, 12, 13
C_DOP, C_DBILL, C_DREC, C_DCL = 16, 17, 18, 19
C_EB, C_DD, C_CHK = 22, 23, 24
C_FLAG = 27

section(fws, 7, "CASH WALK")
line(fws, C_OP, "Opening cash", "€",
     lambda t: f"={A('cash_open')}" if t == 1 else f"={mc(t-1)}{C_CL}", EUR0)
line(fws, C_COL, "Collections (= total billings)", "€", lambda t: f"='P&L'!{mc(t)}{P_BT}", EUR0, font=F_LINK)
line(fws, C_COGS, "Cost of revenue paid", "€", lambda t: f"=-'P&L'!{mc(t)}{P_CT}", EUR0, font=F_LINK)
line(fws, C_OPEX, "Operating expenses paid", "€", lambda t: f"=-'P&L'!{mc(t)}{P_OTT}", EUR0, font=F_LINK)
line(fws, C_NET, "Net cash flow", "€",
     lambda t: f"=SUM({mc(t)}{C_COL}:{mc(t)}{C_OPEX})", EUR0, bold=True, rule=True)
line(fws, C_CL, "Closing cash", "€", lambda t: f"={mc(t)}{C_OP}+{mc(t)}{C_NET}", EUR0, bold=True)

section(fws, 15, "DEFERRED REVENUE ROLLFORWARD  (annual prepayments)")
line(fws, C_DOP, "Opening deferred revenue", "€",
     lambda t: "=0" if t == 1 else f"={mc(t-1)}{C_DCL}", EUR0)
line(fws, C_DBILL, "Annual billings received", "€", lambda t: f"='P&L'!{mc(t)}{P_BA}", EUR0, font=F_LINK)
line(fws, C_DREC, "Recognised into revenue", "€", lambda t: f"=-'P&L'!{mc(t)}{P_RA}", EUR0, font=F_LINK)
line(fws, C_DCL, "Closing deferred revenue", "€",
     lambda t: f"=SUM({mc(t)}{C_DOP}:{mc(t)}{C_DREC})", EUR0, bold=True, rule=True)

section(fws, 21, "INTEGRITY CHECK")
line(fws, C_EB, "EBITDA", "€", lambda t: f"='P&L'!{mc(t)}{P_EB}", EUR0, font=F_LINK)
line(fws, C_DD, "Δ deferred revenue", "€", lambda t: f"={mc(t)}{C_DCL}-{mc(t)}{C_DOP}", EUR0)
line(fws, C_CHK, "Check: net cash − EBITDA − Δdeferred", "€",
     lambda t: f"=ROUND({mc(t)}{C_NET}-{mc(t)}{C_EB}-{mc(t)}{C_DD},6)", EUR2, bold=True)
fws.cell(C_CHK + 1, 1, "Must be zero in every month. Any non-zero cell means the P&L and the cash walk have "
                       "come apart — stop and fix before reading anything else.").font = F_NOTE

fws.cell(C_FLAG - 1, 1, "MEMO").font = F_HDR
line(fws, C_FLAG, "Month flag — net cash flow positive", "#",
     lambda t: f"=IF({mc(t)}{C_NET}>0,{mc(t)}5,9999)", IDX, font=F_NOTE)

# ================================================================ UNIT ECONOMICS
uws = wb.create_sheet("Unit Economics")
uws.sheet_view.showGridLines = False
uws.column_dimensions["A"].width = 44
uws.column_dimensions["B"].width = 14
for col in "CDEFGHIJK":
    uws.column_dimensions[col].width = 12
uws.column_dimensions["L"].width = 74

uws.cell(2, 1, "UNIT ECONOMICS").font = F_TITLE
uws.cell(3, 1, "Steady-state economics on the live scenario, then two grids showing how fast the verdict changes.").font = F_SUB

def u_sec(row, title):
    for c in range(1, 12):
        cell = uws.cell(row, c, title if c == 1 else None)
        cell.fill = FILL_SEC
        if c == 1:
            cell.font = F_SEC

def u_row(row, label, unit, formula, fmt, note="", bold=False, rule=False):
    lc = uws.cell(row, 1, label)
    lc.font = F_LBLB if bold else F_LBL
    uws.cell(row, 2, unit).font = F_NOTE
    uws.cell(row, 2).alignment = Alignment(horizontal="center")
    vc = uws.cell(row, 3, sub(formula))
    vc.font = Font(name="Calibri", size=10, bold=bold, color="000000")
    vc.number_format = fmt
    if rule:
        lc.border = TOP
        vc.border = TOP
        uws.cell(row, 2).border = TOP
    nc = uws.cell(row, 12, note)
    nc.font = F_NOTE
    nc.alignment = Alignment(wrap_text=True, vertical="top")
    uws.row_dimensions[row].height = max(15, 11 * (len(note) // 84 + 1))

u_sec(5, "VALUE OF A CUSTOMER")
u_row(6, "Blended ARPU", "€ / month", "={arpu_blended}", EUR2, "Across monthly and annual plans.")
u_row(7, "Gross margin (month 36, at scale)", "%", f"='P&L'!{LAST}{P_GM}", PCT1,
      "Taken at month 36 so fixed platform cost is spread over the mature base — not the month-1 margin.")
u_row(8, "Gross profit per customer / month", "€", "=C6*C7", EUR2)
u_row(9, "Blended monthly churn", "%", "={churn_blended}", PCT2)
u_row(10, "Average customer lifetime", "months", "={life_months}", NUM1, "1 / blended churn.")
u_row(11, "LTV", "€", "=C8*C10", EUR0, "Gross-profit LTV, undiscounted.", bold=True, rule=True)

u_sec(13, "COST OF A CUSTOMER")
u_row(14, "Paid CAC (funnel maths)", "€", "={paid_cac_theo}", EUR2,
      "CPC / (signup rate x trial-to-paid). What the marginal paid customer costs.")
u_row(15, "Total paid marketing, 36 months", "€", f"=SUM('P&L'!{mc(1)}{P_OM}:{LAST}{P_OM})", EUR0)
u_row(16, "Total new customers, 36 months", "#", f"=SUM(Cohorts!{mc(1)}{R_NEW}:{LAST}{R_NEW})", NUM1)
u_row(17, "Blended CAC", "€", "=IF(C16=0,0,C15/C16)", EUR2,
      "All marketing spend over all new customers, including the organic and referral ones that cost nothing. "
      "Flattering — quote paid CAC to anyone serious.", bold=True, rule=True)

u_sec(19, "VERDICT")
u_row(20, "LTV : CAC — paid", "x", "=IF(C14=0,0,C11/C14)", MULT,
      "The honest one. Below 3.0x, paid acquisition does not fund itself.", bold=True)
u_row(21, "LTV : CAC — blended", "x", "=IF(C17=0,0,C11/C17)", MULT, "Includes free channels.")
u_row(22, "CAC payback — paid", "months", "=IF(C8=0,0,C14/C8)", NUM1,
      "Months of gross profit to repay one paid customer. Must be comfortably below average lifetime (row 10) "
      "or the customer leaves before repaying.", bold=True)
u_row(23, "CAC payback — blended", "months", "=IF(C8=0,0,C17/C8)", NUM1)
u_row(24, "Profitable months after payback", "months", "=C10-C22", NUM1,
      "Average lifetime minus payback. This is the only window in which a customer actually makes money. "
      "Negative means the average customer leaves still owing you their acquisition cost.", bold=True)

# --- Sensitivity 1: LTV:CAC (paid) — churn vs trial conversion
churns = [0.040, 0.045, 0.050, 0.060, 0.070, 0.080, 0.090]
convs  = [0.12, 0.15, 0.18, 0.22, 0.25, 0.28, 0.32]

u_sec(27, "SENSITIVITY — LTV : CAC (paid)")
uws.cell(28, 1, "Blended monthly churn ↓  /  Trial-to-paid →").font = F_HDR
uws.cell(28, 12, "Grid recomputes from the live ARPU, gross margin, CPC and signup rate. "
                 "Green ≥ 3.0x, amber 2.0–3.0x, red < 2.0x — the standard venture bar.").font = F_NOTE
uws.cell(28, 12).alignment = Alignment(wrap_text=True, vertical="top")
uws.row_dimensions[28].height = 30

for j, cv in enumerate(convs):
    c = uws.cell(29, 4 + j, cv)
    c.font = F_IN
    c.number_format = PCT1
    c.alignment = Alignment(horizontal="center")
    c.fill = FILL_BAND
for i, ch in enumerate(churns):
    c = uws.cell(30 + i, 3, ch)
    c.font = F_IN
    c.number_format = PCT1
    c.alignment = Alignment(horizontal="center")
    c.fill = FILL_BAND
    for j, cv in enumerate(convs):
        cl = get_column_letter(4 + j)
        cell = uws.cell(30 + i, 4 + j)
        cell.value = sub(
            f"=IF(OR($C{30+i}=0,{cl}$29=0,{A('cpc')}=0),0,"
            f"(($C$6*$C$7)/$C{30+i})/({A('cpc')}/({A('signup_rate')}*{cl}$29)))"
        )
        cell.font = F_CALC
        cell.number_format = MULT
        cell.alignment = Alignment(horizontal="center")

from openpyxl.formatting.rule import CellIsRule
rng1 = f"D30:{get_column_letter(3+len(convs))}{29+len(churns)}"
uws.conditional_formatting.add(rng1, CellIsRule(operator="greaterThanOrEqual", formula=["3"],
                                                fill=PatternFill("solid", fgColor="C6EFCE")))
uws.conditional_formatting.add(rng1, CellIsRule(operator="between", formula=["2", "3"],
                                                fill=PatternFill("solid", fgColor="FFEB9C")))
uws.conditional_formatting.add(rng1, CellIsRule(operator="lessThan", formula=["2"],
                                                fill=PatternFill("solid", fgColor="FFC7CE")))

# --- Sensitivity 2: CAC payback months — CPC vs trial conversion
cpcs = [0.40, 0.60, 0.80, 1.00, 1.20, 1.50, 2.00]
R2 = 30 + len(churns) + 2

u_sec(R2, "SENSITIVITY — CAC payback (months, paid)")
uws.cell(R2 + 1, 1, "Blended CPC ↓  /  Trial-to-paid →").font = F_HDR
uws.cell(R2 + 1, 12, "Compare every cell against average customer lifetime on row 10. Any cell above it means the "
                     "customer churns before paying you back.").font = F_NOTE
uws.cell(R2 + 1, 12).alignment = Alignment(wrap_text=True, vertical="top")
uws.row_dimensions[R2 + 1].height = 30

for j, cv in enumerate(convs):
    c = uws.cell(R2 + 2, 4 + j, cv)
    c.font = F_IN
    c.number_format = PCT1
    c.alignment = Alignment(horizontal="center")
    c.fill = FILL_BAND
for i, cp in enumerate(cpcs):
    c = uws.cell(R2 + 3 + i, 3, cp)
    c.font = F_IN
    c.number_format = EUR2
    c.alignment = Alignment(horizontal="center")
    c.fill = FILL_BAND
    for j, cv in enumerate(convs):
        cl = get_column_letter(4 + j)
        cell = uws.cell(R2 + 3 + i, 4 + j)
        cell.value = sub(
            f"=IF(OR({cl}${R2+2}=0,$C$6*$C$7=0),0,"
            f"($C{R2+3+i}/({A('signup_rate')}*{cl}${R2+2}))/($C$6*$C$7))"
        )
        cell.font = F_CALC
        cell.number_format = NUM1
        cell.alignment = Alignment(horizontal="center")

rng2 = f"D{R2+3}:{get_column_letter(3+len(convs))}{R2+2+len(cpcs)}"
uws.conditional_formatting.add(rng2, CellIsRule(operator="lessThanOrEqual", formula=["$C$10*0.4"],
                                                fill=PatternFill("solid", fgColor="C6EFCE")))
uws.conditional_formatting.add(rng2, CellIsRule(operator="greaterThan", formula=["$C$10"],
                                                fill=PatternFill("solid", fgColor="FFC7CE")))

# ================================================================ DASHBOARD
dws = wb.create_sheet("Dashboard")
dws.sheet_view.showGridLines = False
dws.column_dimensions["A"].width = 3
dws.column_dimensions["B"].width = 42
dws.column_dimensions["C"].width = 16
dws.column_dimensions["D"].width = 3
dws.column_dimensions["E"].width = 72

dws.cell(2, 2, "DASHBOARD").font = F_TITLE
dws.cell(3, 2, "Headline outputs on the live scenario. Change Assumptions!B5 to switch case.").font = F_SUB

dws.cell(5, 2, "Active scenario").font = F_LBLB
sc = dws.cell(5, 3, "=Assumptions!B6")
sc.font = Font(name="Calibri", size=11, bold=True, color=TEAL)
sc.alignment = Alignment(horizontal="center")

def d_sec(row, title):
    for c in range(2, 6):
        cell = dws.cell(row, c, title if c == 2 else None)
        cell.fill = FILL_SEC
        if c == 2:
            cell.font = F_SEC

def d_row(row, label, formula, fmt, note="", bold=False):
    dws.cell(row, 2, label).font = F_LBLB if bold else F_LBL
    vc = dws.cell(row, 3, sub(formula))
    vc.font = Font(name="Calibri", size=10, bold=bold, color="000000")
    vc.number_format = fmt
    vc.alignment = Alignment(horizontal="center")
    nc = dws.cell(row, 5, note)
    nc.font = F_NOTE
    nc.alignment = Alignment(wrap_text=True, vertical="top")
    dws.row_dimensions[row].height = max(15, 11 * (len(note) // 82 + 1))

d_sec(7, "SCALE AT MONTH 36")
d_row(8, "Active customers", f"=Cohorts!{LAST}{R_ACT}", NUM0)
d_row(9, "MRR", f"='P&L'!{LAST}{P_RT}", EUR0)
d_row(10, "ARR (run-rate)", f"='P&L'!{LAST}{P_ARR}", EUR0, "Month 36 MRR annualised.", bold=True)
d_row(11, "Gross margin", f"='P&L'!{LAST}{P_GM}", PCT1)
d_row(12, "Total revenue, 36 months", f"=SUM('P&L'!{mc(1)}{P_RT}:{LAST}{P_RT})", EUR0)
d_row(13, "Total billings, 36 months", f"=SUM('P&L'!{mc(1)}{P_BT}:{LAST}{P_BT})", EUR0,
      "Higher than revenue by the deferred balance still unearned at month 36.")

d_sec(15, "PROFITABILITY & CASH")
d_row(16, "Cumulative EBITDA, 36 months", f"='P&L'!{LAST}{P_CUM}", EUR0, bold=True)
d_row(17, "First EBITDA-positive month",
      f"=IF(MIN('P&L'!{mc(1)}{P_FLAG}:{LAST}{P_FLAG})>{N},\"none in 36m\",MIN('P&L'!{mc(1)}{P_FLAG}:{LAST}{P_FLAG}))",
      IDX, "Accounting breakeven.", bold=True)
d_row(18, "First cash-flow-positive month",
      f"=IF(MIN('Cash Flow'!{mc(1)}{C_FLAG}:{LAST}{C_FLAG})>{N},\"none in 36m\",MIN('Cash Flow'!{mc(1)}{C_FLAG}:{LAST}{C_FLAG}))",
      IDX, "Earlier than EBITDA breakeven, because annual customers pay before you earn it.")
d_row(19, "Lowest cash balance", f"=MIN('Cash Flow'!{mc(1)}{C_CL}:{LAST}{C_CL})", EUR0)
d_row(20, "Additional funding required",
      f"=MAX(0,-MIN('Cash Flow'!{mc(1)}{C_CL}:{LAST}{C_CL}))", EUR0,
      "Cash needed on top of the opening balance to avoid going negative. Zero means the opening contribution "
      "carries it.", bold=True)
d_row(21, "Closing cash, month 36", f"='Cash Flow'!{LAST}{C_CL}", EUR0)
d_row(22, "Deferred revenue, month 36", f"='Cash Flow'!{LAST}{C_DCL}", EUR0,
      "A liability. It is cash you hold but have not earned — do not read it as profit.")

d_sec(24, "UNIT ECONOMICS")
d_row(25, "LTV", "='Unit Economics'!C11", EUR0)
d_row(26, "Paid CAC", "='Unit Economics'!C14", EUR2)
d_row(27, "LTV : CAC (paid)", "='Unit Economics'!C20", MULT,
      "3.0x is the conventional bar. Below it, growth destroys value.", bold=True)
d_row(28, "CAC payback (paid)", "='Unit Economics'!C22", NUM1, "Months.")
d_row(29, "Average customer lifetime", "='Unit Economics'!C10", NUM1,
      "Months. Payback above this line means the business never repays acquisition.")
d_row(30, "Profitable months after payback", "='Unit Economics'!C24", NUM1,
      "The earning window per customer.")

d_sec(32, "VALUATION")
d_row(33, "NPV of net cash flow, 36 months",
      f"=NPV({A('disc_monthly')},'Cash Flow'!{mc(1)}{C_NET}:{LAST}{C_NET})", EUR0,
      "Discounted at the monthly equivalent of the scenario's annual rate. Excludes terminal value — "
      "a 36-month NPV alone will understate any business that is still compounding at month 36.", bold=True)
d_row(34, "Integrity check (must be 0)",
      f"=SUM('Cash Flow'!{mc(1)}{C_CHK}:{LAST}{C_CHK})", EUR2,
      "Sum of the monthly cash-to-EBITDA reconciliation. Anything other than zero invalidates the sheet.", bold=True)

# --- charts
ch1 = LineChart()
ch1.title = "MRR (€ / month)"
ch1.height = 7.2
ch1.width = 15.5
ch1.style = 2
data1 = Reference(pws, min_col=C0, max_col=C0 + N - 1, min_row=P_RT, max_row=P_RT)
cats = Reference(pws, min_col=C0, max_col=C0 + N - 1, min_row=5, max_row=5)
ch1.add_data(data1, from_rows=True, titles_from_data=False)
ch1.set_categories(cats)
ch1.y_axis.numFmt = '#,##0'
ch1.x_axis.title = "Month"
ch1.legend = None
dws.add_chart(ch1, "G7")

ch2 = LineChart()
ch2.title = "Closing cash (€)"
ch2.height = 7.2
ch2.width = 15.5
ch2.style = 2
data2 = Reference(fws, min_col=C0, max_col=C0 + N - 1, min_row=C_CL, max_row=C_CL)
ch2.add_data(data2, from_rows=True, titles_from_data=False)
ch2.set_categories(cats)
ch2.y_axis.numFmt = '#,##0'
ch2.x_axis.title = "Month"
ch2.legend = None
dws.add_chart(ch2, "G23")

wb.save("/home/claude/Invoice_Chaser_SaaS_Model.xlsx")
print("built")
